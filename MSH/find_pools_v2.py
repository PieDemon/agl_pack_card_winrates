import asyncio
import json
import os
import pandas as pd
from openpyxl import load_workbook
from playwright.async_api import async_playwright

# --- CONFIGURATION ---
MAX_CONCURRENT_WORKERS = 5
INPUT_XLSX = "/Users/jacob.pickering/workspace/agl/bombiest_bombs/pools.xlsx"
TARGET_SHEET = "Sheet93"
DATA_CACHE_FILE = "/Users/jacob.pickering/workspace/agl/bombiest_bombs/downloaded_pools.json"
# ---------------------

def load_existing_cache():
    if os.path.exists(DATA_CACHE_FILE):
        with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return {}
    return {}

def save_cache(cache_data):
    with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, indent=2)

def extract_hyperlink(cell):
    """
    Safely extracts an underlying hyperlink target string from an openpyxl cell object.
    """
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    elif cell.value and str(cell.value).startswith("=HYPERLINK"):
        try:
            # Slices the text between the first set of double quotes in the formula
            return str(cell.value).split('"')[1]
        except IndexError:
            pass
    return None

def parse_sheet93_records():
    if not os.path.exists(INPUT_XLSX):
        print(f"[ CRITICAL ] The file '{INPUT_XLSX}' was not found in this folder.")
        return []

    print(f"Reading target tab '{TARGET_SHEET}' from '{INPUT_XLSX}'...")
    
    # Force openpyxl to isolate our targeted sheet tab structure
    wb = load_workbook(INPUT_XLSX, data_only=False)
    if TARGET_SHEET not in wb.sheetnames:
        print(f"[ CRITICAL ] Tab '{TARGET_SHEET}' does not exist in this workbook.")
        return []
    ws = wb[TARGET_SHEET]

    # Force pandas to parse visible values from the matching sheet tab
    df = pd.read_excel(INPUT_XLSX, sheet_name=TARGET_SHEET)
    
    # Map out header column string keys sequentially
    col_mapping = {col_idx: ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)}
    players_data = []

    # Iterate rows (openpyxl data starts at data row index 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        if row_idx >= len(df): 
            break
            
        player_name = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else None
        if not player_name or player_name.lower() in ['nan', 'none', '']:
            continue

        # Initialize base structural row payload maps
        row_dict = {
            "player_name": player_name,
            "wins": int(float(df.iloc[row_idx, 4])) if pd.notna(df.iloc[row_idx, 4]) else 0,
            "losses": int(float(df.iloc[row_idx, 5])) if pd.notna(df.iloc[row_idx, 5]) else 0,
            "pool_urls": {}
        }

        # Scan columns dynamically to pinpoint STARTING POOL and PACK columns
        for col_idx, cell in enumerate(row, start=1):
            col_name = col_mapping.get(col_idx)
            if not col_name:
                continue
            
            col_name_clean = str(col_name).strip()
            
            # Filter targeted columns explicitly
            if col_name_clean == "STARTING POOL" or col_name_clean.startswith("PACK "):
                link_url = extract_hyperlink(cell)
                if link_url and "sealeddeck.tech" in link_url:
                    row_dict["pool_urls"][col_name_clean] = link_url

        if row_dict["pool_urls"]:
            players_data.append(row_dict)

    return players_data

async def scrape_single_link(browser, player_name, column_label, url, progress_lock, cache_data):
    """
    Downloads cards from a single URL and mounts the output back into the primary cache map.
    """
    context = await browser.new_context()
    page = await context.new_page()
    
    try:
        await page.goto(url, timeout=30000)
        await page.wait_for_selector(".card-node, .deck-builder, img", timeout=60000)
        
        card_elements = await page.query_selector_all("img")
        cards_found = []
        
        for card in card_elements:
            alt_text = await card.get_attribute("alt")
            if alt_text:
                cards_found.append(alt_text.strip())
                
        if cards_found:
            async with progress_lock:
                # Thread-safe insertion back into the local nested player database layout
                cache_data[player_name]["pools"][column_label] = {
                    "url": url,
                    "cards": cards_found
                }
                print(f"[ SAVED ] Scraped {column_label} for {player_name}")
                save_cache(cache_data)
        else:
            print(f"[ WARNING ] Empty layout layout returned for {player_name} -> {column_label}")
            
    except Exception as e:
        print(f"[ ERROR ] Failure resolving pool link ({column_label}) for {player_name}: {e}")
    finally:
        await context.close()

async def worker(queue, browser, progress_lock, cache_data):
    while True:
        try:
            task_payload = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
            
        player_name = task_payload["player_name"]
        column_label = task_payload["column_label"]
        url = task_payload["url"]
        
        await scrape_single_link(browser, player_name, column_label, url, progress_lock, cache_data)
        queue.task_done()

async def main():
    player_records = parse_sheet93_records()
    if not player_records:
        print("[ ABORT ] Found no valid player rows or pool link targets.")
        return

    cache_data = load_existing_cache()
    queue = asyncio.Queue()

    print(f"\nProcessing spreadsheet data metadata details...")
    for record in player_records:
        name = record["player_name"]
        
        # Verify or initialize basic persistent metrics structure for this player profile
        if name not in cache_data:
            cache_data[name] = {
                "wins": record["wins"],
                "losses": record["losses"],
                "pools": {}
            }
        else:
            # Keep win/loss records refreshed if modified on spreadsheet
            cache_data[name]["wins"] = record["wins"]
            cache_data[name]["losses"] = record["losses"]

        # Queue any pool links that are not yet locally downloaded
        for col_label, link_url in record["pool_urls"].items():
            if col_label not in cache_data[name]["pools"]:
                await queue.put({
                    "player_name": name,
                    "column_label": col_label,
                    "url": link_url
                })

    if queue.qsize() == 0:
        print("\nAll player records and individual packs are fully downloaded and cached!")
        return

    print(f"Syncing engine online. Discovered {len(player_records)} player profiles.")
    print(f"Queueing {queue.qsize()} individual pack URLs across threads for processing...")
    
    progress_lock = asyncio.Lock()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        workers = [
            asyncio.create_task(worker(queue, browser, progress_lock, cache_data))
            for _ in range(MAX_CONCURRENT_WORKERS)
        ]
        await queue.join()
        for w in workers:
            w.cancel()
        await browser.close()
        
    print(f"\nAll operations finished. Cache state saved to '{DATA_CACHE_FILE}'.")

if __name__ == "__main__":
    asyncio.run(main())

