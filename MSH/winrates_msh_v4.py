import json
import os
import csv
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURATION ---
INPUT_JSON = "downloaded_pools.json"  
INPUT_XLSX = "pools.xlsx"
WHITELIST_TXT = "msh_cards.txt"
OUTPUT_CSV = "msh_card_post_acquisition_stats.csv"

TARGET_SHEET = "Sheet93"
MATCHES_SHEET = "Matches"
# ---------------------

def clean_player_name(raw_name):
    if not raw_name or pd.isna(raw_name):
        return ""
    name_str = str(raw_name).strip()
    if " - " in name_str:
        return name_str.split(" - ")[0].strip()
    if "-" in name_str:
        return name_str.split("-")[0].strip()
    return name_str

def load_msh_whitelist():
    if not os.path.exists(WHITELIST_TXT):
        print(f"[ CRITICAL ] Whitelist file '{WHITELIST_TXT}' not found.")
        return set()
    msh_set = set()
    with open(WHITELIST_TXT, "r", encoding="utf-8") as f:
        for line in f:
            clean_name = line.strip().lower()
            if clean_name:
                msh_set.add(clean_name)
    print(f"[ SUCCESS ] Loaded {len(msh_set)} unique Marvel card definitions.")
    return msh_set

def compute_pure_match_history_timeline():
    if not os.path.exists(INPUT_XLSX):
        return {}, {}

    print(f"Re-computing pure player records from '{MATCHES_SHEET}'...")
    df_matches = pd.read_excel(INPUT_XLSX, sheet_name=MATCHES_SHEET)
    player_stats = {}

    for _, row in df_matches.iterrows():
        winner = clean_player_name(row.get("Your Name", ""))
        loser = clean_player_name(row.get("Loser Name", ""))

        for p in [winner, loser]:
            if p and p.lower() not in ['nan', 'none', ''] and p not in player_stats:
                player_stats[p] = {"wins": 0, "losses": 0, "loss_milestones": []}

        if winner in player_stats:
            player_stats[winner]["wins"] += 1
            
        if loser in player_stats:
            current_w = player_stats[loser]["wins"]
            current_l = player_stats[loser]["losses"] + 1
            player_stats[loser]["loss_milestones"].append(f"{current_w}-{current_l}")
            player_stats[loser]["losses"] += 1

    timeline_milestones = {k: v["loss_milestones"] for k, v in player_stats.items()}
    final_true_records = {k: (v["wins"], v["losses"]) for k, v in player_stats.items()}
    return timeline_milestones, final_true_records

def build_sheet93_pack_order_map():
    if not os.path.exists(INPUT_XLSX):
        return {}, []
    wb = load_workbook(INPUT_XLSX, data_only=True)
    ws = wb[TARGET_SHEET]
    df = pd.read_excel(INPUT_XLSX, sheet_name=TARGET_SHEET)
    
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers_clean = [str(h).strip() for h in headers if h]
    pool_columns_in_order = [h for h in headers_clean if h == "STARTING POOL" or h.startswith("PACK ")]
    
    player_pack_maps = {}
    for row_idx in range(len(df)):
        player_name = str(df.iloc[row_idx, 0]).strip()
        if not player_name or player_name.lower() in ['nan', 'none', '']:
            continue
            
        player_present_packs = []
        for col_name in pool_columns_in_order:
            col_idx_df = df.columns.get_loc(col_name) if col_name in df.columns else None
            if col_idx_df is not None:
                cell_value = df.iloc[row_idx, col_idx_df]
                if pd.notna(cell_value) and str(cell_value).strip() != "":
                    player_present_packs.append(col_name)
        player_pack_maps[player_name] = player_present_packs
        
    return player_pack_maps, pool_columns_in_order

def main():
    if not os.path.exists(INPUT_JSON):
        print(f"[ CRITICAL ] Scraped cache '{INPUT_JSON}' not found.")
        return

    msh_whitelist = load_msh_whitelist()
    timeline_milestones, final_true_records = compute_pure_match_history_timeline()
    player_pack_maps, pack_chronology = build_sheet93_pack_order_map()

    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        scraped_data = json.load(f)

    card_performance = {}
    print("Processing intervals and applying duplicate multipliers...")

    for player_name, profile in scraped_data.items():
        if player_name not in final_true_records:
            continue  
            
        final_true_w, final_true_l = final_true_records[player_name]
        pools = profile.get("pools", {})
        ordered_packs = player_pack_maps.get(player_name, [])
        player_losses_list = timeline_milestones.get(player_name, [])

        pack_records = {}
        last_w, last_l = 0, 0
        l_idx = 0
        
        for col_label in ordered_packs:
            if col_label not in pools:
                continue
            if col_label == "STARTING POOL" or col_label == "PACK 1":
                pack_records[col_label] = (0, 0)
            else:
                l_idx += 1
                if (l_idx - 1) < len(player_losses_list):
                    w_str, l_str = player_losses_list[l_idx - 1].split('-')
                    pack_records[col_label] = (int(w_str), int(l_str))
                    last_w, last_l = int(w_str), int(l_str)
                else:
                    pack_records[col_label] = (last_w, last_l)

        all_discovered_msh_cards = set()
        for col_label in ordered_packs:
            if col_label in pools:
                for c in pools[col_label].get("cards", []):
                    if c.strip().lower() in msh_whitelist:
                        all_discovered_msh_cards.add(c.strip())

        for card in all_discovered_msh_cards:
            running_copies_held = 0
            
            for i, col_label in enumerate(ordered_packs):
                if col_label not in pools:
                    continue
                    
                pack_cards = pools[col_label].get("cards", [])
                opened_in_this_pack = sum(1 for c in pack_cards if c.strip() == card)
                running_copies_held += opened_in_this_pack

                if running_copies_held == 0:
                    continue  

                w_start, l_start = pack_records[col_label]

                next_pack_found = False
                for next_col in ordered_packs[i+1:]:
                    if next_col in pools:
                        w_end, l_end = pack_records[next_col]
                        next_pack_found = True
                        break
                if not next_pack_found:
                    w_end, l_end = final_true_w, final_true_l

                era_wins = w_end - w_start
                era_losses = l_end - l_start

                if era_wins < 0 or era_losses < 0:
                    continue

                weighted_wins = era_wins * running_copies_held
                weighted_losses = era_losses * running_copies_held

                if card not in card_performance:
                    card_performance[card] = {"wins": 0, "losses": 0, "total_post_matches": 0, "pools": 0}

                card_performance[card]["wins"] += weighted_wins
                card_performance[card]["losses"] += weighted_losses
                card_performance[card]["total_post_matches"] += (weighted_wins + weighted_losses)
                if i == 0 or (card_performance[card]["pools"] == 0): 
                    card_performance[card]["pools"] += 1

    final_rankings = []

    for card_name, stats in card_performance.items():
        weighted_matches = stats["total_post_matches"]
        if weighted_matches == 0:
            continue
            
        weighted_wins = stats["wins"]
        raw_weighted_winrate = (weighted_wins / weighted_matches) * 100
        
        final_rankings.append({
            "Card Name": card_name,
            "Weighted Winrate %": round(raw_weighted_winrate, 2),
            "Weighted Record": f"{weighted_wins}-{stats['losses']}",
            "Total Weighted Matches": weighted_matches,
            "Pools Opened In": stats["pools"]
        })

    # Sort descending by raw, duplicate-weighted winrate
    sorted_rankings = sorted(final_rankings, key=lambda x: x["Weighted Winrate %"], reverse=True)

    print(f"\n--- PURIFIED MARVEL SUPER HEROES SCOREBOARD (RAW WEIGHTED) ---")
    print(f"{'RANK':<5} | {'CARD NAME':<50} | {'WEIGHTED WR':<14} | {'WEIGHTED RECORD':<16} | {'POOLS'}")
    print("-" * 103)
    for rank, item in enumerate(sorted_rankings[:50], 1):
        raw_str = f"{item['Weighted Winrate %']}%"
        print(f"{rank:<5} | {item['Card Name']:<50} | {raw_str:<14} | {item['Weighted Record']:<16} | {item['Pools Opened In']}")

    with open(OUTPUT_CSV, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Card Name", "Weighted Winrate %", "Weighted Record", "Total Weighted Matches", "Pools Opened In"
        ])
        writer.writeheader()
        writer.writerows(sorted_rankings)

    print(f"\n[ COMPLETE ] Raw duplicate-weighted analysis saved to '{OUTPUT_CSV}'")

if __name__ == "__main__":
    main()

