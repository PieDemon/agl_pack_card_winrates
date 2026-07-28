import json
import os
import re
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURATION ---
INPUT_JSON = "./SOS/sos_pools.json"
INPUT_XLSX = "./SOS/sos.xlsx"
POOLS_SHEET = "ExtractedPools"
# ---------------------

def clean_player_name(raw_name):
    if not raw_name or pd.isna(raw_name):
        return ""
    return str(raw_name).strip()

def extract_hyperlink_target(cell):
    """Selects the longest string match to skip formula boilerplate fragments."""
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
        
    cell_str = str(cell.value or "").strip()
    links = re.findall(r'https?://[^\s"\')]+', cell_str)
    if links:
        return max(links, key=len).strip()
    return None

def main():
    if not os.path.exists(INPUT_JSON) or not os.path.exists(INPUT_XLSX):
        print("[ ERROR ] Make sure both the json and xlsx files exist in the target paths.")
        return

    # 1. Load your current JSON containing all your successfully scraped packs
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        json_db = json.load(f)

    # 2. Open workbook in formula mode to parse out the starting pool links
    print(f"Reading target sheet '{POOLS_SHEET}' to isolate starting pool columns...")
    wb = load_workbook(INPUT_XLSX, data_only=False)
    ws = wb[POOLS_SHEET]
    
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers_clean = [str(h).strip().upper() if h else "" for h in headers]
    
    id_col_idx = next((i + 1 for i, h in enumerate(headers_clean) if "IDENTIFICATION" in h), None)
    start_pool_col_idx = next((i + 1 for i, h in enumerate(headers_clean) if "STARTING POOL" in h), None)

    if not id_col_idx or not start_pool_col_idx:
        print("[ CRITICAL ] Could not map the 'Identification' or 'Starting Pool' headers.")
        return

    patched_count = 0

    # 3. Iterate rows and patch ONLY the STARTING POOL URL entries in memory
    for row_idx in range(2, ws.max_row + 1):
        player_id = clean_player_name(ws.cell(row=row_idx, column=id_col_idx).value)
        if not player_id or player_id not in json_db:
            continue

        cell = ws.cell(row=row_idx, column=start_pool_col_idx)
        correct_url = extract_hyperlink_target(cell)

        if correct_url and "sealeddeck.tech" in correct_url:
            # Safely navigate to the specific key without resetting or emptying 'cards' arrays
            if "STARTING POOL" in json_db[player_id]["pools"]:
                current_url = json_db[player_id]["pools"]["STARTING POOL"].get("url", "")
                
                # Only update if the URL is currently broken or empty
                if len(current_url) <= 26 or "STARTING POOL" not in json_db[player_id]["pools"]:
                    json_db[player_id]["pools"]["STARTING POOL"]["url"] = correct_url
                    patched_count += 1

    # 4. Save the patched database back to the exact same file
    with open(INPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(json_db, f, indent=2)

    print(f"\n[ SUCCESS ] Successfully patched {patched_count} STARTING POOL URLs.")
    print(f"All previously scraped card lists for Packs 1-10 are fully preserved.")

if __name__ == "__main__":
    main()

