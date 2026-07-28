import os
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURATION ---
INPUT_XLSX = "pools.xlsx"
TARGET_SHEET = "Sheet93"
# ---------------------

def inspect_pack1_raw_cells():
    if not os.path.exists(INPUT_XLSX):
        print(f"[ ERROR ] '{INPUT_XLSX}' not found.")
        return

    print("--- OPENING WORKBOOK IN FORMULA MODE (data_only=False) ---")
    wb_formula = load_workbook(INPUT_XLSX, data_only=False)
    ws_formula = wb_formula[TARGET_SHEET]

    print("\n--- OPENING WORKBOOK IN VALUE MODE (data_only=True) ---")
    wb_value = load_workbook(INPUT_XLSX, data_only=True)
    ws_value = wb_value[TARGET_SHEET]

    # Map the headers out precisely
    headers = [ws_formula.cell(row=1, column=c).value for c in range(1, ws_formula.max_column + 1)]
    
    pack1_idx = None
    for idx, name in enumerate(headers, start=1):
        if name and str(name).strip().upper().replace(" ", "") == "PACK1":
            pack1_idx = idx
            break

    if not pack1_idx:
        print(f"[ CRITICAL ] Could not identify column 'PACK 1'. Checked headers: {headers}")
        return

    print(f"Target Column 'PACK 1' located at physical spreadsheet Index: {pack1_idx}\n")
    print(f"{'PLAYER NAME':<20} | {'RAW FORMULA/VALUE property':<45} | {'NATIVE LINK TARGET'}")
    print("-" * 90)

    # Scan the first 15 player data rows to inspect patterns
    count = 0
    for row_idx in range(2, ws_formula.max_row + 1):
        player_name = str(ws_formula.cell(row=row_idx, column=1).value or "").strip()
        if not player_name or player_name.lower() in ['nan', 'none']:
            continue
            
        count += 1
        if count > 15:
            break

        cell_f = ws_formula.cell(row=row_idx, column=pack1_idx)
        cell_v = ws_value.cell(row=row_idx, column=pack1_idx)

        # Extract underlying openpyxl metadata parameters
        raw_val = cell_f.value
        evaluated_val = cell_v.value
        native_link = cell_f.hyperlink.target if cell_f.hyperlink else "None"

        print(f"{player_name:<20} | Type: {type(raw_val).__name__:<8} | Val: {str(raw_val)[:30]:<30} | Link: {native_link}")

if __name__ == "__main__":
    inspect_pack1_raw_cells()

