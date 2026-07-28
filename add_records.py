import json
import os
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURATION ---
INPUT_XLSX = "pools.xlsx"
INPUT_JSON = "downloaded_pools.json"
OUTPUT_JSON = "updated_pools.json"

TARGET_SHEET = "Sheet93"
MATCHES_SHEET = "Matches"
# ---------------------

def clean_player_name(raw_name):
    """
    Cleans a timeline string name (e.g., 'Nick A - Nistar#14127' or 'Nick A') 
    to match the short name used in Sheet93 ('Nick A').
    """
    if not raw_name or pd.isna(raw_name):
        return ""
    
    name_str = str(raw_name).strip()
    
    if " - " in name_str:
        name_str = name_str.split(" - ")[0].strip()
    elif "-" in name_str:
        name_str = name_str.split("-")[0].strip()
        
    return name_str

def compute_player_records_timeline():
    """
    Parses the 'Matches' sheet to reconstruct a running tally of wins and losses 
    for every player chronologically, tracking what their record was AT THE MOMENT 
    of receiving a consecutive loss.
    """
    if not os.path.exists(INPUT_XLSX):
        print(f"[ ERROR ] '{INPUT_XLSX}' not found. Cannot compute match history.")
        return {}

    print(f"Parsing '{MATCHES_SHEET}' timeline to reconstruct match history...")
    try:
        df_matches = pd.read_excel(INPUT_XLSX, sheet_name=MATCHES_SHEET)
    except ValueError:
        print(f"[ CRITICAL ] Tab '{MATCHES_SHEET}' not found in your Excel file.")
        return {}

    history = {}

    for _, row in df_matches.iterrows():
        winner = clean_player_name(row.get("Your Name", ""))
        loser = clean_player_name(row.get("Loser Name", ""))

        for player in [winner, loser]:
            if player and player.lower() not in ['nan', 'none', ''] and player not in history:
                history[player] = {"wins": 0, "losses": 0, "loss_history_records": []}

        if winner in history:
            history[winner]["wins"] += 1

        if loser in history:
            # FIX: Increment loss count by 1 to include the loss that triggered this pack
            current_record_str = f"{history[loser]['wins']}-{history[loser]['losses'] + 1}"
            history[loser]["loss_history_records"].append(current_record_str)
            history[loser]["losses"] += 1

    return history

def build_sheet93_pack_order_map():
    """
    Scans Sheet93 to see exactly which columns exist and mapping order, 
    ensuring we parse pack sequencing identical to how they appear on your spreadsheet.
    """
    if not os.path.exists(INPUT_XLSX):
        return {}

    wb = load_workbook(INPUT_XLSX, data_only=True)
    if TARGET_SHEET not in wb.sheetnames:
        print(f"[ CRITICAL ] Tab '{TARGET_SHEET}' missing from Excel file.")
        return {}
        
    ws = wb[TARGET_SHEET]
    df = pd.read_excel(INPUT_XLSX, sheet_name=TARGET_SHEET)
    
    headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    headers_clean = [str(h).strip() for h in headers if h]

    pool_columns_in_order = [h for h in headers_clean if h == "STARTING POOL" or h.startswith("PACK ")]
    
    player_pack_maps = {}
    
    for row_idx in range(len(df)):
        player_name = str(df.iloc[row_idx, 0]).strip()
        if not player_name or player_name.lower() in ['nan', 'none', '']:
            continue
            
        player_present_packs = []
        for col_name in pool_columns_in_order:
            col_idx_df = df.columns.get_loc(col_name) if col_name in df.columns else None
            if col_idx_df is not None:
                cell_value = df.iloc[row_idx, col_idx_df]
                if pd.notna(cell_value) and str(cell_value).strip() != "":
                    player_present_packs.append(col_name)
                    
        player_pack_maps[player_name] = player_present_packs
        
    return player_pack_maps

def main():
    if not os.path.exists(INPUT_JSON):
        print(f"[ CRITICAL ] Scraped database file '{INPUT_JSON}' was not found in this folder.")
        return

    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        scraped_data = json.load(f)

    # 1. Generate running histories from cleaned match timeline rows
    timeline_history = compute_player_records_timeline()

    # 2. Figure out the columns ordering and populated cells from Sheet93
    player_pack_maps = build_sheet93_pack_order_map()

    print(f"Injecting historical timeline data and resolving entropy gaps...")
    updated_count = 0
    entropy_resolved_count = 0

    # Loop through the players currently existing inside your json file
    for player_name, data in scraped_data.items():
        if "pools" not in data:
            continue
            
        ordered_packs = player_pack_maps.get(player_name, [])
        player_history = timeline_history.get(player_name)
        
        # Keep a running fallback state tracker for this user to calculate entropy offsets
        last_known_wins = 0
        last_known_losses = 0
        
        pack_index = 0
        
        # Traverse the packs in physical layout order (left to right)
        for col_label in ordered_packs:
            if col_label in data["pools"]:
                
                if col_label == "STARTING POOL" or col_label == "PACK 1":
                    record_at_receipt = "0-0"
                    last_known_wins = 0
                    last_known_losses = 0
                else:
                    pack_index += 1  # Tracks the match-loss index count (PACK 2 = index 0)
                    
                    # Try to pull the clean match-result log state
                    if player_history and (pack_index - 1) < len(player_history["loss_history_records"]):
                        record_at_receipt = player_history["loss_history_records"][pack_index - 1]
                        
                        # Sync our tracker with the confirmed match logs
                        try:
                            w, l = map(int, record_at_receipt.split('-'))
                            last_known_wins = w
                            last_known_losses = l
                        except ValueError:
                            pass
                    else:
                        # FIX: Handle an entropy loss! Take the last record and increment loss by 1 before recording string
                        last_known_losses += 1
                        record_at_receipt = f"{last_known_wins}-{last_known_losses}"
                        entropy_resolved_count += 1
                
                # Update or append into existing dictionary key paths safely
                pack_data = data["pools"][col_label]
                if isinstance(pack_data, list):
                    data["pools"][col_label] = {
                        "record_at_receipt": record_at_receipt,
                        "cards": pack_data
                    }
                else:
                    data["pools"][col_label]["record_at_receipt"] = record_at_receipt
                
                updated_count += 1

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(scraped_data, f, indent=2)

    print(f"\nSuccess! Modified {updated_count} individual pack elements seamlessly.")
    if entropy_resolved_count > 0:
        print(f"[ FIXED ] Successfully resolved {entropy_resolved_count} entropy gaps by incrementing running player losses.")
    print(f"Your fully enriched database has been written to: '{OUTPUT_JSON}'")

if __name__ == "__main__":
    main()

