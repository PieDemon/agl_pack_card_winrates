import json
import os
import re
import sys
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURATION DEFAULT ---
DEFAULT_INPUT_XLSX = "./SOS/sos.xlsx"
POOLS_SHEET = "ExtractedPools"
MATCHES_SHEET = "Matches"
# -----------------------------

def get_io_paths():
    if len(sys.argv) > 1:
        input_xlsx = sys.argv[1].strip()
    else:
        input_xlsx = DEFAULT_INPUT_XLSX
    base_path, _ = os.path.splitext(input_xlsx)
    return input_xlsx, f"{base_path}_pools.json"

def clean_player_name(raw_name):
    if not raw_name or pd.isna(raw_name):
        return ""
    return str(raw_name).strip()

def extract_hyperlink_target(cell):
    """
    Extracts a URL string out of native Excel hyperlink properties, or runs
    a regex filter that selects the longest full-length URL match inside 
    complex nested formulas to skip over baseline text fragments.
    """
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
        
    cell_str = str(cell.value or "").strip()
    
    # Locate all available HTTP web strings inside the cell text
    links = re.findall(r'https?://[^\s"\')]+', cell_str)
    if links:
        # FIX: Sort by string length descending to select the complete, valid URL 
        # (e.g., matching 'https://sealeddeck.tech/tnFRwCeGTe' over 'https://sealeddeck.tech/')
        longest_url = max(links, key=len)
        return longest_url.strip()
        
    return None

def compute_pure_match_history_timeline(input_xlsx):
    if not os.path.exists(input_xlsx):
        print(f"[ ERROR ] '{input_xlsx}' file not found.")
        return {}, {}

    print(f"Parsing '{MATCHES_SHEET}' timeline to reconstruct pure player records...")
    df_matches = pd.read_excel(input_xlsx, sheet_name=MATCHES_SHEET)
    player_stats = {}

    for _, row in df_matches.iterrows():
        winner = clean_player_name(row.get("Your Name", ""))
        loser = clean_player_name(row.get("Loser Name", ""))

        for player in [winner, loser]:
            if player and player.lower() not in ['nan', 'none', ''] and player not in player_stats:
                player_stats[player] = {"wins": 0, "losses": 0, "loss_milestones": []}

        if winner in player_stats:
            player_stats[winner]["wins"] += 1
            
        if loser in player_stats:
            current_w = player_stats[loser]["wins"]
            current_l = player_stats[loser]["losses"] + 1
            player_stats[loser]["loss_milestones"].append(f"{current_w}-{current_l}")
            player_stats[loser]["losses"] += 1

    timeline_milestones = {k: v["loss_milestones"] for k, v in player_stats.items()}
    final_true_records = {k: (v["wins"], v["losses"]) for k, v in player_stats.items()}
    return timeline_milestones, final_true_records

def main():
    input_xlsx, output_json = get_io_paths()

    if not os.path.exists(input_xlsx):
        print(f"[ CRITICAL ] The target source file '{input_xlsx}' was not found.")
        return

    # 1. Reconstruct pure records completely independent of entropy columns
    timeline_milestones, final_true_records = compute_pure_match_history_timeline(input_xlsx)

    print(f"\nOpening workbook to extract pack links from '{POOLS_SHEET}'...")
    wb = load_workbook(input_xlsx, data_only=False)
    if POOLS_SHEET not in wb.sheetnames:
        print(f"[ CRITICAL ] Tab '{POOLS_SHEET}' does not exist in your Excel file.")
        return
    ws = wb[POOLS_SHEET]
    
    df_pools = pd.read_excel(input_xlsx, sheet_name=POOLS_SHEET)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers_clean = [str(h).strip().upper() if h else "" for h in headers]
    
    id_col_idx = next((i + 1 for i, h in enumerate(headers_clean) if "IDENTIFICATION" in h), None)
    if not id_col_idx:
        print("[ CRITICAL ] Could not find an 'Identification' column header in row 1.")
        return
        
    print(f"[ DEBUG ] Mapping 'Identification' keys from Excel column index: {id_col_idx}")
    sos_json_db = {}
    processed_packs_count = 0

    # 2. Iterate row by row through players
    for row_idx in range(2, ws.max_row + 1):
        player_id = clean_player_name(ws.cell(row=row_idx, column=id_col_idx).value)
        if not player_id or player_id.lower() in ['nan', 'none', '']:
            continue

        final_w, final_l = final_true_records.get(player_id, (0, 0))
        player_losses_list = timeline_milestones.get(player_id, [])

        sos_json_db[player_id] = {
            "wins": final_w,
            "losses": final_l,
            "pools": {}
        }

        loss_pack_index = 0
        last_known_w, last_known_l = 0, 0

        # Scan columns sequentially to catch STARTING POOL and PACK columns
        for col_idx in range(1, ws.max_column + 1):
            col_name = headers[col_idx - 1]
            if not col_name:
                continue
                
            col_name_clean = str(col_name).strip().upper()
            
            if col_name_clean == "STARTING POOL" or col_name_clean.startswith("PACKS ") or col_name_clean.startswith("PACK "):
                cell = ws.cell(row=row_idx, column=col_idx)
                link_url = extract_hyperlink_target(cell)
                
                if link_url and "sealeddeck.tech" in link_url:
                    if col_name_clean == "STARTING POOL" or col_name_clean == "PACKS 1" or col_name_clean == "PACK 1":
                        w_at_receipt, l_at_receipt = 0, 0
                        last_known_w, last_known_l = 0, 0
                    else:
                        loss_pack_index += 1  
                        
                        if (loss_pack_index - 1) < len(player_losses_list):
                            record_str = player_losses_list[loss_pack_index - 1]
                            w_at_receipt, l_at_receipt = map(int, record_str.split('-'))
                            last_known_w, last_known_l = w_at_receipt, l_at_receipt
                        else:
                            w_at_receipt, l_at_receipt = last_known_w, last_known_l

                    standard_label = col_name_clean.replace("PACKS", "PACK")

                    sos_json_db[player_id]["pools"][standard_label] = {
                        "url": link_url,
                        "record_at_receipt": f"{w_at_receipt}-{l_at_receipt}",
                        "cards": []  
                    }
                    processed_packs_count += 1

    out_dir = os.path.dirname(output_json)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(sos_json_db, f, indent=2)

    print(f"\n[ COMPLETE ] Extracted {len(sos_json_db)} player profiles.")
    print(f"Successfully compiled {processed_packs_count} verified pack nodes into: '{output_json}'")

if __name__ == "__main__":
    main()

